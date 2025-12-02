import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import re
import io
import os
import time
import base64
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 0. KONFIGURACJA
# ==========================================

st.set_page_config(
    page_title="Stopy&Dystans", 
    layout="wide", 
    page_icon="📦",
    initial_sidebar_state="expanded"
)

# --- USTAWIENIA INTRA ---
PLIK_WIDEO = "logo.mp4"
CZAS_TRWANIA_INTRA = 10 

def get_base64_video(video_path):
    with open(video_path, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()

st.markdown("""
<style>
    .stApp { background-color: #000000; color: #ffffff; }
    [data-testid="stSidebar"] { background-color: #050505; border-right: 1px solid #333; }
    h1, h2, h3, h4, h5, h6, p, label, .stMarkdown, div { color: #ffffff !important; }
    div.stMetric { background-color: #111111 !important; border: 1px solid #333 !important; }
    #intro-container {
        position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
        background-color: black; z-index: 999999;
        display: flex; justify_content: center; align-items: center; flex-direction: column;
    }
    header[data-testid="stHeader"] { background-color: transparent; }
</style>
""", unsafe_allow_html=True)

if 'intro_played' not in st.session_state:
    st.session_state['intro_played'] = False

if not st.session_state['intro_played'] and os.path.exists(PLIK_WIDEO):
    intro_placeholder = st.empty()
    video_b64 = get_base64_video(PLIK_WIDEO)
    intro_html = f"""
    <div id="intro-container">
        <video autoplay loop playsinline style="width: 50%; max-width: 600px;">
            <source src="data:video/mp4;base64,{video_b64}" type="video/mp4">
        </video>
    </div>
    """
    intro_placeholder.markdown(intro_html, unsafe_allow_html=True)
    time.sleep(CZAS_TRWANIA_INTRA)
    intro_placeholder.empty()
    st.session_state['intro_played'] = True


# ==========================================
# 1. FUNKCJA SKANUJĄCA MAPĘ WIZUALNĄ
# ==========================================

def parse_visual_map(uploaded_file):
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    coords = {}
    
    header_row_idx = None
    rack_cols = {} 
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        found_racks = 0
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, (int, float)) and val > 100:
                found_racks += 1
        if found_racks > 2:
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if isinstance(val, (int, float)):
                    rack_cols[c_idx] = int(val)
            break
            
    if not header_row_idx:
        return None, "Nie znaleziono rzędu z numerami regałów (szukałem liczb > 100)."

    for row in ws.iter_rows(min_row=header_row_idx+1, values_only=False):
        for cell in row:
            if cell.column in rack_cols:
                val = cell.value
                if val and isinstance(val, str):
                    match = re.search(r'^(\d+)', str(val).strip())
                    if match:
                        try:
                            col_num = int(match.group(1))
                            rack_num = rack_cols[cell.column]
                            x = cell.column
                            y = cell.row
                            coords[(rack_num, col_num)] = (x, y)
                        except: pass
    return coords, None


# ==========================================
# 2. SIDEBAR
# ==========================================

with st.sidebar:
    if os.path.exists(PLIK_WIDEO):
        video_b64 = get_base64_video(PLIK_WIDEO)
        logo_html = f"""
        <div style="display: flex; justify-content: center; margin-bottom: 20px;">
            <video autoplay loop muted playsinline style="width: 120px; border-radius: 8px;">
                <source src="data:video/mp4;base64,{video_b64}" type="video/mp4">
            </video>
        </div>
        """
        st.markdown(logo_html, unsafe_allow_html=True)

    st.markdown("---")
    st.header("📂 Dane wejściowe")
    uploaded_file = st.file_uploader("1. Wgraj analizę (analiza.xlsx)", type=["xlsx"])
    uploaded_map = st.file_uploader("2. Wgraj mapę magazynu (opcjonalne - wizualną)", type=["xlsx"])

# ==========================================
# 3. LOGIKA BIZNESOWA (SKRYPT VD)
# ==========================================

# --- NAPRAWIONE STAŁE (TU BYŁ BŁĄD) ---
FRONT = "front"
P034  = "034"
P057  = "057"
P058  = "058"

A_FRONT_TO_034 = 72.5     
B_034_TO_057   = 66.0     
C_058_TO_062   = 11.7     
CROSS_LANES    = 5.4      
START_STOP     = 9.0      
WRONG_ENTRY_PENALTY = 100000.0 
ENTRY_SIDE_SOFT_PENALTY = 0.0
U_TURN_PENALTY = 200.0         
NEIGHBOR_MIDDLE_PENALTY = 1.0 
BRIDGE_PENALTY = 0.0

TRYBY_EXT_ALL = {680, 690}
TRYB_670_EXT  = {1022,1023,1024,1025,1026}

START_STOP_MAP = {
    602:(826,825), 601:(826,825), 610:(826,825),
    722:(826,825), 622:(826,825), 620:(826,825),
    630:(859,860), 640:(859,860), 641:(859,860),
    650:(859,860), 655:(859,860), 660:(859,860),
    670:(1020,1019), 680:(1020,1019), 690:(1020,1019),
}
AXIS_602_LANE   = 800
SS_TO_AXIS_602  = 70.2     
LENGTH_602_LOOP = 146.0    
NEEDED = {
 "Numer misji":["Numer misji","numer misji","misja","nr misji","id misji"],
 "Tryb Pracy":["Tryb Pracy","tryb pracy","tryb"],
 "SKU":["SKU","sku"],
 "numer lini":["numer lini","numer linii","linia","line"],
 "Regal":["Regal","Regał","regał","alejka","regal"],
 "Kolumna":["Kolumna","kolumna","kol"],
 "Poziom":["Poziom","poziom","level"],
 "miejsce":["miejsce","slot","miejsce pobrania"],
}

# --- FUNKCJE POMOCNICZE ---
def tryb_ma_rozszerzenie(tryb, lane): return tryb in TRYBY_EXT_ALL or (tryb == 670 and lane in TRYB_670_EXT)
def pair_key(reg): return int(reg) if int(reg) % 2 == 1 else int(reg) - 1
def normalize_rack(reg_num): return reg_num - 20 if reg_num >= 901 else reg_num
def pair_gaps(a,b): return abs(normalize_rack(int(a)) - normalize_rack(int(b))) // 2
def _to_tryb_int(v): m=re.search(r'(\d+)',str(v)); return int(m.group(1)) if m else None
def _try_int(x): m=re.search(r'(-?\d+)',str(x)); return int(m.group(1)) if m and pd.notna(x) else None
def _auto_map(df):
    cols=list(df.columns); low=[str(c).strip().lower() for c in cols]
    m={k: next((cols[i] for i,h in enumerate(low) if any(n.lower() in h for n in v)), None) for k,v in NEEDED.items()}
    if not m["Regal"] or not m["Kolumna"]: raise ValueError("Brakuje kolumn Regał/Kolumna")
    return m
def fmt_loc(reg,col,poziom,miejsce):
    k=str(int(col)).zfill(3); p="00"
    if isinstance(poziom,str): mm=re.search(r'(\d+)',poziom); p=mm.group(1).zfill(2) if mm else "00"
    elif pd.notna(poziom): p=str(int(poziom)).zfill(2)
    m=str(miejsce).strip() if pd.notna(miejsce) else "C"
    return f"{int(reg)}.{k}{p}{m}"

def get_m(state, nC, ext):
    if state==0: return 0.0
    if state==2: return A_FRONT_TO_034
    if nC and ext: return A_FRONT_TO_034 + B_034_TO_057 + C_058_TO_062
    return A_FRONT_TO_034 + B_034_TO_057

def _lane_distance(entry, exit, nA, nB, nC, ext):
    s = 0.0 if entry==FRONT else (A_FRONT_TO_034 if entry==P034 else A_FRONT_TO_034+B_034_TO_057)
    t = 0.0 if exit==FRONT else (A_FRONT_TO_034 if exit==P034 else A_FRONT_TO_034+B_034_TO_057)
    req = 0.0
    if nA: req = max(req, A_FRONT_TO_034)
    if nB: req = max(req, A_FRONT_TO_034 + B_034_TO_057)
    if nC and ext: req = max(req, A_FRONT_TO_034 + B_034_TO_057 + C_058_TO_062)
    if req==0.0: return abs(t-s)
    return round(req + min(abs(s)+abs(t-req), abs(s-req)+abs(t)), 1)

def calc_aisle_cost_real(state_in, state_out, data):
    cost = 10**9; nA, nB, nC, ext = data['nA'], data['nB'], data['nC'], data['ext']
    anchor_back = (P058 if (ext and nC) else P057)
    if state_in==0 and state_out==0: cost = _lane_distance(FRONT, FRONT, nA, nB, nC, ext) + (U_TURN_PENALTY if nA else 0)
    elif (state_in==0 and state_out==1) or (state_in==1 and state_out==0): cost = _lane_distance(FRONT, anchor_back, nA, nB, nC, ext)
    elif state_in==1 and state_out==1: cost = _lane_distance(anchor_back, anchor_back, nA, nB, nC, ext) + (U_TURN_PENALTY if nB else 0)
    elif (state_in==0 and state_out==2) or (state_in==2 and state_out==0): cost = A_FRONT_TO_034 if not (nB or nC) else 10**9
    elif (state_in==1 and state_out==2) or (state_in==2 and state_out==1): cost = B_034_TO_057 + (C_058_TO_062 if ext and nC else 0) if not nA else 10**9
    elif state_in==2 and state_out==2: cost = 0.0 if not (nA or nB or nC) else 10**9
    return cost

def simulate_blocks(grp, tryb, start_pair):
    g=grp.copy(); det=[]; total=0.0
    blocks=[]; curp=None; cols=[]
    for _,r in g.iterrows():
        p=pair_key(int(r["Regal"])); c=int(r["Kolumna"])
        if curp is None: curp=p
        if p!=curp:
            if cols: blocks.append((curp,cols)); cols=[]
            curp=p
        cols.append(c)
    if cols: blocks.append((curp,cols))
    if not blocks: return 0.0, start_pair, []

    hop = pair_gaps(start_pair, blocks[0][0]) * CROSS_LANES
    if hop>0: det.append((f"Front: {start_pair} -> {blocks[0][0]}", hop)); total+=hop

    info=[]
    for p,cl in blocks:
        ci=[int(c) for c in cl]
        info.append({'p':p, 'nA':any(c<=33 for c in ci), 'nB':any(c>=35 for c in ci), 'nC':any(c>=59 for c in ci), 'ext':tryb_ma_rozszerzenie(tryb,p), 'first_col':ci[0]})

    n=len(info); dp=[[10**9]*3 for _ in range(n)]; prev=[[None]*3 for _ in range(n)]
    curr=info[0]
    for i in range(3): dp[0][i] = calc_aisle_cost_real(0, i, curr)

    for i in range(n-1):
        for s_out in range(3):
            if dp[i][s_out] >= 10**9: continue
            curr, next_blk = info[i], info[i+1]
            gaps = pair_gaps(curr['p'], next_blk['p'])
            for s_in in range(3):
                cross = gaps * CROSS_LANES
                if gaps==1 and s_out==2 and s_in==2: cross += NEIGHBOR_MIDDLE_PENALTY
                if gaps>=2 and s_out!=s_in: 
                    m_out = get_m(s_out, curr['nC'], curr['ext'])
                    m_in = get_m(s_in, next_blk['nC'], next_blk['ext'])
                    cross += abs(m_out-m_in) + BRIDGE_PENALTY
                
                # Kara za wejście
                pref = 0 if next_blk['first_col']<=34 else 1
                pen = 0.0
                if s_in != pref: pen += WRONG_ENTRY_PENALTY if gaps>=2 else ENTRY_SIDE_SOFT_PENALTY
                
                trans = cross + pen
                for s_tgt in range(3):
                    new_cost = dp[i][s_out] + trans + calc_aisle_cost_real(s_in, s_tgt, next_blk)
                    if new_cost < dp[i+1][s_tgt]: dp[i+1][s_tgt] = new_cost; prev[i+1][s_tgt] = (s_out, 0, s_in) # Method 0 dummy

    best_end=0; min_val=dp[-1][0]
    for s in range(1,3): 
        if dp[-1][s] < min_val: min_val=dp[-1][s]; best_end=s
    
    path=[]; curr_s=best_end
    for i in range(n-1, 0, -1):
        p_s, _, in_s = prev[i][curr_s]
        path.append((p_s, 0, in_s, curr_s)); curr_s=p_s
    path.reverse(); path.insert(0, (None,0,0,curr_s))

    accum_debug = total
    for i in range(n):
        curr=info[i]; _,_,s_in,s_out=path[i]
        cost = calc_aisle_cost_real(s_in, s_out, curr)
        lbl_in = "A" if s_in==0 else "034" if s_in==2 else "B"
        lbl_out = "A" if s_out==0 else "034" if s_out==2 else "B"
        det.append((f"Alejka {curr['p']}: {lbl_in}->{lbl_out}", cost))
        accum_debug += cost
        if i < n-1:
            gaps = pair_gaps(curr['p'], info[i+1]['p'])
            cross_c = gaps * CROSS_LANES
            det.append((f"Przejazd do {info[i+1]['p']}", cross_c))
            accum_debug += cross_c

    last = info[-1]; ss_t = START_STOP_MAP.get(tryb); ss_p = pair_key(min(ss_t)) if ss_t else last['p']
    ret_gap = pair_gaps(last['p'], ss_p)
    if ret_gap>0: det.append((f"Powrót do {ss_p}", ret_gap*CROSS_LANES)); accum_debug+=ret_gap*CROSS_LANES
    
    return round(accum_debug,1), ss_p, det

def simulate_602(grp, last_p, ss_p):
    d = []
    hop = pair_gaps(last_p, pair_key(AXIS_602_LANE)) * CROSS_LANES
    if hop>0: d.append(("Dojazd do 602", hop))
    d.append(("602 Pętla", LENGTH_602_LOOP))
    return round(hop+LENGTH_602_LOOP, 1), pair_key(AXIS_602_LANE), d

@st.cache_data
def process_data(uploaded_file):
    try: raw=pd.read_excel(uploaded_file,header=0,dtype=object); m=_auto_map(raw)
    except: 
        try: raw=pd.read_excel(uploaded_file,header=1,dtype=object); m=_auto_map(raw)
        except Exception as e: return None, None, str(e)
    df=raw.rename(columns={m[k]:k for k in m if m[k]})
    if "Numer misji" in df.columns: df["Numer misji"]=df["Numer misji"].astype(str).str.strip()
    if "Tryb Pracy" in df.columns: df["Tryb Pracy"]=df["Tryb Pracy"].astype(str).str.strip()
    if "numer lini" not in df.columns: df["numer lini"]=(df.groupby("Numer misji").cumcount()+1)*10
    df["Regal"]=df["Regal"].apply(_try_int); df["Kolumna"]=df["Kolumna"].apply(_try_int)
    df=df.dropna(subset=["Regal","Kolumna"]).astype({"Regal":int, "Kolumna":int})
    df["__sort__"]=pd.to_numeric(df["numer lini"], errors="coerce").fillna(9e9)
    
    rows_d=[]; rows_s=[]
    for (misja,tryb), grp in df.groupby(["Numer misji","Tryb Pracy"]):
        grp=grp.sort_values("__sort__", kind="mergesort")
        ss_t = START_STOP_MAP.get(_to_tryb_int(tryb))
        ss_p = pair_key(min(ss_t)) if ss_t else pair_key(AXIS_602_LANE)
        
        last_p = ss_p; dist=START_STOP*2
        
        # Split by blocks
        blocks=[]; cur_t=None; buf=[]
        for _,r in grp.iterrows():
            t="602" if r["Regal"] in (600,601) else "aisle"
            if cur_t is None: cur_t=t
            if t!=cur_t: blocks.append((cur_t, pd.DataFrame(buf))); buf=[]; cur_t=t
            buf.append(r)
        if buf: blocks.append((cur_t, pd.DataFrame(buf)))
        
        for t, bdf in blocks:
            if t=="602": d, last_p, _ = simulate_602(bdf, last_p, ss_p)
            else: d, last_p, _ = simulate_blocks(bdf, _to_tryb_int(tryb), last_p)
            dist += d
        
        dist += pair_gaps(last_p, ss_p)*CROSS_LANES
        
        rows_s.append({"Numer misji":misja, "Tryb Pracy":tryb, "Dystans (m)":round(dist,1), "Ilość stopów":len(grp.drop_duplicates(["Regal","Kolumna"]))})
        for _,r in grp.iterrows():
            d=r.to_dict(); d.update({"Dystans (m)":round(dist,1)}); rows_d.append(d)
            
    return pd.DataFrame(rows_d), pd.DataFrame(rows_s), None

def generate_excel_download(df_det, df_sum):
    output = io.BytesIO()
    wb = Workbook(); ws1=wb.active; ws1.title="Szczegóły"
    for r in dataframe_to_rows(df_det,index=False,header=True): ws1.append(r)
    ws2=wb.create_sheet("Podsumowanie")
    for r in dataframe_to_rows(df_sum,index=False,header=True): ws2.append(r)
    wb.save(output); return output.getvalue()

# ==========================================
# 4. INTERFEJS
# ==========================================

st.title("📦 VD: Analiza Stopy i Dystans")
st.markdown("**Automatyczna analiza ścieżek kompletacyjnych i wizualizacja tras**")

# Ładowanie mapy
map_coords = {}
if uploaded_map:
    # 1. Próba: Format 4 kolumn (Regal, Kolumna, X, Y)
    try:
        df_map = pd.read_excel(uploaded_map)
        req = ["Regal", "Kolumna", "X", "Y"]
        if all(c in df_map.columns for c in req):
            for _, r in df_map.iterrows():
                map_coords[(int(r['Regal']), int(r['Kolumna']))] = (float(r['X']), float(r['Y']))
            st.sidebar.success(f"Wczytano mapę prostą: {len(map_coords)} pkt")
    except: pass

    # 2. Próba: Format wizualny (Skaner)
    if not map_coords:
        coords, err = parse_visual_map(uploaded_map)
        if coords:
            map_coords = coords
            st.sidebar.success(f"Zeskanowano mapę wizualną: {len(map_coords)} pkt!")
        elif err:
            st.sidebar.warning(f"Problem z mapą: {err}")

if uploaded_file:
    df_det, df_sum, error = process_data(uploaded_file)
    if error: st.error(error)
    else:
        st.sidebar.success("Gotowe!")
        k1,k2,k3 = st.columns(3)
        k1.metric("Dystans Total", f"{df_sum['Dystans (m)'].sum()/1000:.2f} km")
        k2.metric("Misje", df_sum['Numer misji'].nunique())
        k3.metric("Śr. Dystans", f"{df_sum['Dystans (m)'].mean():.1f} m")
        
        st.subheader("🗺️ Wizualizacja Trasy")
        sel_m = st.selectbox("Wybierz misję:", df_det["Numer misji"].unique())
        m_data = df_det[df_det["Numer misji"] == sel_m].sort_values("__sort__")
        
        def get_xy(r, c):
            if (r,c) in map_coords: return map_coords[(r,c)]
            return (pair_key(r) + (0.2 if r%2==0 else -0.2), c)
        
        path_x = []; path_y = []; lbls = []
        for _, row in m_data.iterrows():
            x, y = get_xy(int(row["Regal"]), int(row["Kolumna"]))
            path_x.append(x); path_y.append(y); lbls.append(fmt_loc(row["Regal"],row["Kolumna"],row.get("Poziom"),row.get("miejsce")))
            
        fig = go.Figure()
        
        # Tło
        if map_coords:
            mx = [v[0] for v in map_coords.values()]; my = [v[1] for v in map_coords.values()]
            # Odwracamy Y, żeby góra Excela była górą mapy (opcjonalne, zależy od preferencji)
            # Tutaj zostawiam standardowo.
            fig.add_trace(go.Scatter(x=mx, y=my, mode='markers', marker=dict(color='#333', size=4), name='Regały', hoverinfo='none'))
            
        # Ścieżka
        fig.add_trace(go.Scatter(x=path_x, y=path_y, mode='lines+markers', line=dict(color='#0066cc', width=3), marker=dict(color='orange', size=8), text=lbls, name='Trasa'))
        
        # Start/Stop
        if path_x:
            fig.add_trace(go.Scatter(x=[path_x[0]], y=[path_y[0]], mode='markers', marker=dict(color='green', size=12, symbol='triangle-right'), name='Start'))
            fig.add_trace(go.Scatter(x=[path_x[-1]], y=[path_y[-1]], mode='markers', marker=dict(color='red', size=12, symbol='square'), name='Koniec'))

        fig.update_layout(height=700, paper_bgcolor='#111', plot_bgcolor='#111', font=dict(color='white'),
                          xaxis=dict(showgrid=False, zeroline=False, visible=False), 
                          yaxis=dict(showgrid=False, zeroline=False, visible=False, autorange="reversed")) # Odwracamy Y żeby pasowało do Excela (1. rząd na górze)
        st.plotly_chart(fig, use_container_width=True)
        
        st.download_button("Pobierz Raport", generate_excel_download(df_det, df_sum), "Raport.xlsx")
